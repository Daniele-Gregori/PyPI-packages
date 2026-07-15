"""Core spreadsheet utilities.

Python port of the Wolfram Language package ``DanieleGregori`SpreadsheetToolkit```:

- ``import_all``        <-> ``ImportAll``
- ``index_to_position`` <-> ``IndexToPosition``
- ``position_to_index`` <-> ``PositionToIndex``
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Union

import openpyxl

try:
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:  # pragma: no cover - very old openpyxl
    ArrayFormula = ()

# ---------------------------------------------------------------------------
# ImportAll
# ---------------------------------------------------------------------------

# Mirrors ResourceFunction["ImportOnce"]: repeated imports of the same
# (unmodified) file return the cached result.
_import_cache: dict = {}


def import_all(file: Union[str, Path]) -> tuple:
    """Import sheet names, data and formulas from a spreadsheet file.

    Mirrors ``ImportAll[file]``, i.e. the Wolfram Language import
    ``Import[file, {{"Sheets", "Data", "Formulas"}}]`` cached with
    ``ImportOnce``.

    Parameters
    ----------
    file : str or Path
        Path to an ``.xlsx`` file.

    Returns
    -------
    tuple
        ``(sheets, data, formulas)`` where ``sheets`` is the list of sheet
        names, ``data`` is a list (one entry per sheet) of 2D lists of cell
        values, and ``formulas`` is a list of same-shaped 2D lists holding the
        formula string (without the leading ``=``) for formula cells and
        ``""`` elsewhere. Empty data cells are also ``""``.

    Examples
    --------
    >>> sheets, data, formulas = import_all("book.xlsx")  # doctest: +SKIP
    """
    path = Path(file).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"No such spreadsheet file: {str(path)!r}")

    key = (str(path), path.stat().st_mtime_ns)
    cached = _import_cache.get(key)
    if cached is not None:
        return cached

    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    try:
        sheets = list(wb_formulas.sheetnames)
        data = []
        formulas = []
        for name in sheets:
            grid_values, grid_formulas = _sheet_grids(wb_values[name], wb_formulas[name])
            data.append(grid_values)
            formulas.append(grid_formulas)
    finally:
        wb_values.close()
        wb_formulas.close()

    result = (sheets, data, formulas)
    _import_cache[key] = result
    return result


def _sheet_grids(ws_values, ws_formulas) -> tuple:
    """Build the (data, formulas) grids of one worksheet.

    Following the Wolfram "Data"/"Formulas" import elements: empty cells and
    the data entries of formula cells become ``""``; the formulas grid holds
    ``""`` for non-formula cells and the formula string without ``=``.
    """
    max_row = ws_formulas.max_row
    max_col = ws_formulas.max_column
    grid_values: list = []
    grid_formulas: list = []
    rows_v = ws_values.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
    rows_f = ws_formulas.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
    for row_v, row_f in zip(rows_v, rows_f):
        values_row: list = []
        formulas_row: list = []
        for cell_v, cell_f in zip(row_v, row_f):
            raw_formula = cell_f.value
            if isinstance(raw_formula, ArrayFormula):
                raw_formula = raw_formula.text or ""
            if isinstance(raw_formula, str) and raw_formula.startswith("="):
                formulas_row.append(raw_formula[1:])
            else:
                formulas_row.append("")
            value = cell_v.value
            values_row.append("" if value is None else value)
        grid_values.append(values_row)
        grid_formulas.append(formulas_row)
    return grid_values, grid_formulas


# ---------------------------------------------------------------------------
# IndexToPosition
# ---------------------------------------------------------------------------

_INDEX_RE = re.compile(r"\$?([A-Za-z]+)\$?(\d+)")


def index_to_position(index: str) -> tuple:
    """Convert a spreadsheet cell reference to a ``(row, column)`` position.

    Mirrors ``IndexToPosition[cellpatt]``: the row is the number in the
    reference, the column is the base-26 value of the letters; ``$`` markers
    of absolute references are ignored.

    Parameters
    ----------
    index : str
        A cell reference such as ``"A1"``, ``"AA12"`` or ``"$B$10"``.

    Returns
    -------
    tuple
        ``(row, column)`` with 1-based indices.

    Examples
    --------
    >>> index_to_position("A1")
    (1, 1)
    >>> index_to_position("C5")
    (5, 3)
    >>> index_to_position("AA1")
    (1, 27)
    >>> index_to_position("$B$10")
    (10, 2)
    """
    if not isinstance(index, str):
        raise TypeError(f"Cell reference must be a string, got {type(index).__name__}")
    match = _INDEX_RE.fullmatch(index)
    if match is None:
        raise ValueError(f"Invalid cell reference: {index!r}")
    letters, row = match.group(1), int(match.group(2))
    column = 0
    for char in letters.upper():
        column = column * 26 + (ord(char) - ord("A") + 1)
    return (row, column)


# ---------------------------------------------------------------------------
# PositionToIndex
# ---------------------------------------------------------------------------


def position_to_index(position) -> str:
    """Convert a ``(row, column)`` position to a spreadsheet cell reference.

    Mirrors ``PositionToIndex[{r, c}]``: the column number is written in
    bijective base 26 with letters ``A``-``Z`` and the row is appended.

    Parameters
    ----------
    position : sequence of two ints
        ``(row, column)`` with 1-based indices.

    Returns
    -------
    str
        A cell reference such as ``"A1"`` or ``"AA12"``.

    Examples
    --------
    >>> position_to_index((1, 1))
    'A1'
    >>> position_to_index((5, 3))
    'C5'
    >>> position_to_index((1, 27))
    'AA1'
    """
    row, column = position
    if not isinstance(row, int) or not isinstance(column, int):
        raise TypeError(f"Position must be a pair of ints, got {position!r}")
    if row < 1:
        raise ValueError(f"Row must be >= 1, got {row}")
    if column < 1:
        raise ValueError(f"Column must be >= 1, got {column}")
    letters = []
    while column > 0:
        column, remainder = divmod(column - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters)) + str(row)


# ---------------------------------------------------------------------------
# Backward-compatible aliases (0.1.x names)
# ---------------------------------------------------------------------------

spreadsheet_index_to_position = index_to_position
position_to_spreadsheet_index = position_to_index


# ---------------------------------------------------------------------------
# Backward-compatible import helpers (0.1.x names, derived from import_all)
# ---------------------------------------------------------------------------


def _select_grid(grids: list, sheets: list, sheet: Union[str, int, None]) -> list:
    """Pick one per-sheet grid by name or 1-based index (``None`` = first)."""
    if sheet is None:
        index = 1
    elif isinstance(sheet, int):
        if not 1 <= sheet <= len(sheets):
            raise ValueError(
                f"Sheet index {sheet} out of range (workbook has {len(sheets)} sheets)"
            )
        index = sheet
    else:
        if sheet not in sheets:
            raise ValueError(f"Sheet {sheet!r} not found (available: {sheets})")
        index = sheets.index(sheet) + 1
    return grids[index - 1]


def _slice_grid(
    grid: list,
    rows: Union[tuple, None],
    columns: Union[tuple, None],
) -> list:
    """Slice a grid by inclusive 1-based (first, last) row/column ranges,
    returning fresh lists so the ``import_all`` cache stays untouched."""
    if rows is not None:
        grid = grid[rows[0] - 1 : rows[1]]
    if columns is not None:
        return [list(row[columns[0] - 1 : columns[1]]) for row in grid]
    return [list(row) for row in grid]


def import_sheets(file: Union[str, Path]) -> list:
    """Return the list of sheet names in a spreadsheet file.

    Backward-compatible helper (0.1.x name), equivalent to
    ``import_all(file)[0]``.
    """
    sheets, _, _ = import_all(file)
    return list(sheets)


def import_cells(
    file: Union[str, Path],
    sheet: Union[str, int, None] = None,
    rows: Union[tuple, None] = None,
    columns: Union[tuple, None] = None,
) -> list:
    """Import cell values from one sheet of a spreadsheet.

    Backward-compatible helper (0.1.x name), consistent with
    :func:`import_all`: empty cells (and formula cells without a cached
    value) import as ``""``.

    Parameters
    ----------
    file : str or Path
        Path to an ``.xlsx`` file.
    sheet : str, int, or None
        Sheet name (str) or 1-based index (int). ``None`` uses the first sheet.
    rows : tuple or None
        ``(first_row, last_row)`` inclusive, 1-based. ``None`` reads all rows.
    columns : tuple or None
        ``(first_col, last_col)`` inclusive, 1-based. ``None`` reads all columns.

    Returns
    -------
    list
        A 2D list of cell values (one inner list per row).
    """
    sheets, data, _ = import_all(file)
    return _slice_grid(_select_grid(data, sheets, sheet), rows, columns)


def import_formulas(
    file: Union[str, Path],
    sheet: Union[str, int, None] = None,
    rows: Union[tuple, None] = None,
    columns: Union[tuple, None] = None,
) -> list:
    """Import cell formulas from one sheet of a spreadsheet.

    Backward-compatible helper (0.1.x name), consistent with
    :func:`import_all`: formula cells hold the formula string without the
    leading ``=`` and all other cells hold ``""``.

    Parameters are as in :func:`import_cells`.
    """
    sheets, _, formulas = import_all(file)
    return _slice_grid(_select_grid(formulas, sheets, sheet), rows, columns)
