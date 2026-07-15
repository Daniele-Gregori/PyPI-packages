"""Tests for the core utilities: ``import_all``, ``index_to_position`` and
``position_to_index``.

Includes the coverage of the 0.1.x test suite, ported to the current API
(see the class docstrings for the behavior changes).
"""

from pathlib import Path

import openpyxl
import pytest

from spreadsheet_toolkit import import_all, index_to_position, position_to_index


# ---------------------------------------------------------------------------
# Fixtures (ported from the 0.1.x suite)
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a sample .xlsx with two sheets, values, and formulas."""
    p = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()

    # Sheet1 (active)
    ws1 = wb.active
    ws1.title = "Data"
    ws1["A1"] = 10
    ws1["B1"] = 20
    ws1["C1"] = "=A1+B1"
    ws1["A2"] = "hello"
    ws1["B2"] = "world"
    ws1["C2"] = None
    ws1["A3"] = 3.14
    ws1["B3"] = True
    ws1["C3"] = "=SUM(A1:B1)"

    # Sheet2
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "total"
    ws2["B1"] = "=Data!A1+Data!B1"

    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def empty_xlsx(tmp_path):
    """Create an .xlsx with a single empty sheet."""
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """Create an .xlsx with five sheets."""
    p = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Alpha"
    for name in ["Beta", "Gamma", "Delta", "Epsilon"]:
        wb.create_sheet(name)
    wb.save(p)
    wb.close()
    return p


# ---------------------------------------------------------------------------
# index_to_position
# ---------------------------------------------------------------------------


class TestIndexToPosition:
    """Ports ``TestSpreadsheetIndexToPosition`` (0.1.x). Change: absolute
    references like ``$A$1`` are now valid, following the Wolfram
    ``IndexToPosition``, which ignores ``$`` markers."""

    @pytest.mark.parametrize(
        ("index", "position"),
        [
            ("A1", (1, 1)),
            ("B1", (1, 2)),
            ("B3", (3, 2)),
            ("C5", (5, 3)),
            ("Z1", (1, 26)),
            ("AA1", (1, 27)),
            ("AZ1", (1, 52)),
            ("BA1", (1, 53)),
            ("ZZ1", (1, 702)),
            ("AAA1", (1, 703)),
            ("XFD1", (1, 16384)),
            ("D100", (100, 4)),
            ("E999", (999, 5)),
            ("D142", (142, 4)),
        ],
    )
    def test_plain_references(self, index, position):
        assert index_to_position(index) == position

    def test_large_row(self):
        assert index_to_position("A1048576") == (1048576, 1)

    def test_case_insensitive(self):
        assert index_to_position("a1") == (1, 1)
        assert index_to_position("aA3") == (3, 27)
        assert index_to_position("Bc10") == (10, 55)

    @pytest.mark.parametrize(
        ("index", "position"),
        [
            ("$A$1", (1, 1)),
            ("$B$10", (10, 2)),
            ("$B10", (10, 2)),
            ("B$10", (10, 2)),
            ("$AA$12", (12, 27)),
        ],
    )
    def test_absolute_references(self, index, position):
        assert index_to_position(index) == position

    @pytest.mark.parametrize(
        "bad",
        ["", "A", "ZZZ", "123", "1A", "A1B", "A1!", "A 1", "A1:B2"],
    )
    def test_invalid_reference_raises(self, bad):
        with pytest.raises(ValueError):
            index_to_position(bad)

    def test_non_string_raises(self):
        with pytest.raises(TypeError):
            index_to_position(123)


# ---------------------------------------------------------------------------
# position_to_index
# ---------------------------------------------------------------------------


class TestPositionToIndex:
    """Ports ``TestPositionToSpreadsheetIndex`` (0.1.x). Change: the position
    is now a single ``(row, column)`` pair, as in the Wolfram
    ``PositionToIndex[{r, c}]``."""

    @pytest.mark.parametrize(
        ("position", "index"),
        [
            ((1, 1), "A1"),
            ((1, 2), "B1"),
            ((3, 2), "B3"),
            ((5, 3), "C5"),
            ((1, 26), "Z1"),
            ((1, 27), "AA1"),
            ((1, 52), "AZ1"),
            ((1, 53), "BA1"),
            ((1, 702), "ZZ1"),
            ((1, 703), "AAA1"),
            ((1, 16384), "XFD1"),
            ((142, 4), "D142"),
        ],
    )
    def test_positions(self, position, index):
        assert position_to_index(position) == index

    def test_large_row(self):
        assert position_to_index((1048576, 1)) == "A1048576"

    @pytest.mark.parametrize("bad", [(0, 1), (-1, 1), (1, 0), (1, -5)])
    def test_out_of_range_raises(self, bad):
        with pytest.raises(ValueError):
            position_to_index(bad)

    def test_non_integer_raises(self):
        with pytest.raises(TypeError):
            position_to_index((1.5, "A"))


# ---------------------------------------------------------------------------
# Round-trip
# ---------------------------------------------------------------------------


class TestRoundTrip:
    @pytest.mark.parametrize(
        "index",
        ["A1", "Z99", "AA1", "AZ52", "BA100", "ZZ1", "ABC123", "XFD1048576"],
    )
    def test_index_roundtrip(self, index):
        assert position_to_index(index_to_position(index)) == index.upper()

    @pytest.mark.parametrize(
        ("row", "col"),
        [(1, 1), (7, 12), (1, 26), (1, 27), (1, 702), (1, 703), (999, 16384)],
    )
    def test_position_roundtrip(self, row, col):
        assert index_to_position(position_to_index((row, col))) == (row, col)


# ---------------------------------------------------------------------------
# import_all
# ---------------------------------------------------------------------------


class TestImportAll:
    """Ports the ``import_sheets`` / ``import_cells`` / ``import_formulas``
    coverage (0.1.x) to ``import_all``, which follows the Wolfram
    ``Import[file, {{"Sheets", "Data", "Formulas"}}]`` conventions: empty
    cells import as ``""``, the formulas grid has ``""`` for non-formula
    cells, and formula strings lose their leading ``=``."""

    # -- sheets (was TestImportSheets) --

    def test_two_sheets(self, sample_xlsx):
        sheets, _, _ = import_all(sample_xlsx)
        assert sheets == ["Data", "Summary"]

    def test_single_sheet(self, empty_xlsx):
        sheets, _, _ = import_all(empty_xlsx)
        assert sheets == ["Empty"]

    def test_five_sheets(self, multi_sheet_xlsx):
        sheets, _, _ = import_all(multi_sheet_xlsx)
        assert sheets == ["Alpha", "Beta", "Gamma", "Delta", "Epsilon"]

    # -- data (was TestImportCells) --

    def test_all_cells_first_sheet(self, sample_xlsx):
        _, data, _ = import_all(sample_xlsx)
        assert len(data[0]) == 3
        assert data[0][0][0] == 10
        assert data[0][0][1] == 20
        assert data[0][1][0] == "hello"
        assert data[0][1][1] == "world"
        assert data[0][2][0] == 3.14
        assert data[0][2][1] is True

    def test_formula_cells_have_cached_value_or_empty(self, sample_xlsx):
        _, data, _ = import_all(sample_xlsx)
        # C1 = "=A1+B1": freshly created files carry no cached value,
        # which imports as "" (was None in 0.1.x)
        assert data[0][0][2] in ("", 30)

    def test_second_sheet_data(self, sample_xlsx):
        _, data, _ = import_all(sample_xlsx)
        assert data[1][0][0] == "total"

    def test_empty_cell_is_empty_string(self, sample_xlsx):
        _, data, _ = import_all(sample_xlsx)
        assert data[0][1][2] == ""  # C2 (was None in 0.1.x)

    def test_empty_sheet(self, empty_xlsx):
        _, data, formulas = import_all(empty_xlsx)
        assert data == [[[""]]]  # was [[None]] in 0.1.x
        assert formulas == [[[""]]]

    def test_accepts_string_path(self, sample_xlsx):
        sheets, data, _ = import_all(str(sample_xlsx))
        assert sheets == ["Data", "Summary"]
        assert data[0][0][0] == 10

    def test_accepts_path_object(self, sample_xlsx):
        _, data, _ = import_all(Path(sample_xlsx))
        assert data[0][0][0] == 10

    # -- formulas (was TestImportFormulas) --

    def test_formula_cells_return_formulas(self, sample_xlsx):
        _, _, formulas = import_all(sample_xlsx)
        # leading "=" is stripped (was kept in 0.1.x)
        assert formulas[0][0][2] == "A1+B1"
        assert formulas[0][2][2] == "SUM(A1:B1)"

    def test_non_formula_cells_are_empty_strings(self, sample_xlsx):
        _, _, formulas = import_all(sample_xlsx)
        # was the cell value in 0.1.x
        assert formulas[0][0][0] == ""
        assert formulas[0][1][0] == ""

    def test_second_sheet_formula(self, sample_xlsx):
        _, _, formulas = import_all(sample_xlsx)
        assert formulas[1][0][0] == ""
        assert formulas[1][0][1] == "Data!A1+Data!B1"

    # -- fixture-file checks --

    def test_fixture_data_values(self, file1):
        _, data, _ = import_all(file1)
        assert data[0][0][0] == 10
        assert data[0][4][0] == "hello"

    def test_fixture_sheet_names(self, file4):
        sheets, _, _ = import_all(file4)
        assert sheets == ["Input", "Summary"]

    def test_grids_are_rectangular_and_aligned(self, file4):
        _, data, formulas = import_all(file4)
        for grid_values, grid_formulas in zip(data, formulas):
            widths = {len(row) for row in grid_values}
            assert len(widths) == 1
            assert len(grid_values) == len(grid_formulas)
            assert all(
                len(rv) == len(rf) for rv, rf in zip(grid_values, grid_formulas)
            )

    def test_cross_sheet_formula_kept_verbatim(self, file4):
        _, _, formulas = import_all(file4)
        assert formulas[1][2][1] == "Input!B14"

    def test_import_once_caching(self, file1):
        assert import_all(file1) is import_all(file1)

    def test_missing_file_raises(self):
        with pytest.raises(FileNotFoundError):
            import_all("no_such_file.xlsx")
