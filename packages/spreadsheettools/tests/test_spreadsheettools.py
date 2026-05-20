"""Tests for spreadsheettools."""

import pytest
import openpyxl
from pathlib import Path
from spreadsheettools import (
    spreadsheet_index_to_position,
    position_to_spreadsheet_index,
    import_sheets,
    import_cells,
    import_formulas,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a sample .xlsx with two sheets, values, and formulas."""
    p = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()

    # Sheet1 (active)
    ws1 = wb.active
    ws1.title = "Data"
    ws1["A1"] = 10
    ws1["B1"] = 20
    ws1["C1"] = "=A1+B1"
    ws1["A2"] = "hello"
    ws1["B2"] = "world"
    ws1["C2"] = None
    ws1["A3"] = 3.14
    ws1["B3"] = True
    ws1["C3"] = "=SUM(A1:B1)"

    # Sheet2
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "total"
    ws2["B1"] = "=Data!A1+Data!B1"

    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def empty_xlsx(tmp_path):
    """Create an .xlsx with a single empty sheet."""
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """Create an .xlsx with five sheets."""
    p = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Alpha"
    for name in ["Beta", "Gamma", "Delta", "Epsilon"]:
        wb.create_sheet(name)
    wb.save(p)
    wb.close()
    return p


# ===========================================================================
# spreadsheet_index_to_position
# ===========================================================================

class TestSpreadsheetIndexToPosition:
    def test_a1(self):
        assert spreadsheet_index_to_position("A1") == (1, 1)

    def test_single_letter_columns(self):
        assert spreadsheet_index_to_position("B1") == (1, 2)
        assert spreadsheet_index_to_position("C5") == (5, 3)
        assert spreadsheet_index_to_position("Z1") == (1, 26)

    def test_double_letter_columns(self):
        assert spreadsheet_index_to_position("AA1") == (1, 27)
        assert spreadsheet_index_to_position("AZ1") == (1, 52)
        assert spreadsheet_index_to_position("BA1") == (1, 53)
        assert spreadsheet_index_to_position("ZZ1") == (1, 702)

    def test_triple_letter_columns(self):
        assert spreadsheet_index_to_position("AAA1") == (1, 703)
        assert spreadsheet_index_to_position("XFD1") == (1, 16384)

    def test_case_insensitive(self):
        assert spreadsheet_index_to_position("a1") == (1, 1)
        assert spreadsheet_index_to_position("aA3") == (3, 27)
        assert spreadsheet_index_to_position("Bc10") == (10, 55)

    def test_large_row(self):
        assert spreadsheet_index_to_position("A1048576") == (1048576, 1)

    def test_various_rows(self):
        assert spreadsheet_index_to_position("D100") == (100, 4)
        assert spreadsheet_index_to_position("E999") == (999, 5)

    def test_invalid_empty(self):
        with pytest.raises(ValueError):
            spreadsheet_index_to_position("")

    def test_invalid_no_row(self):
        with pytest.raises(ValueError):
            spreadsheet_index_to_position("A")

    def test_invalid_no_col(self):
        with pytest.raises(ValueError):
            spreadsheet_index_to_position("123")

    def test_invalid_special_chars(self):
        with pytest.raises(ValueError):
            spreadsheet_index_to_position("A1!")

    def test_invalid_spaces(self):
        with pytest.raises(ValueError):
            spreadsheet_index_to_position("A 1")

    def test_invalid_dollar_sign(self):
        with pytest.raises(ValueError):
            spreadsheet_index_to_position("$A$1")


# ===========================================================================
# position_to_spreadsheet_index
# ===========================================================================

class TestPositionToSpreadsheetIndex:
    def test_a1(self):
        assert position_to_spreadsheet_index(1, 1) == "A1"

    def test_single_letter_columns(self):
        assert position_to_spreadsheet_index(1, 2) == "B1"
        assert position_to_spreadsheet_index(5, 3) == "C5"
        assert position_to_spreadsheet_index(1, 26) == "Z1"

    def test_double_letter_columns(self):
        assert position_to_spreadsheet_index(1, 27) == "AA1"
        assert position_to_spreadsheet_index(1, 52) == "AZ1"
        assert position_to_spreadsheet_index(1, 53) == "BA1"
        assert position_to_spreadsheet_index(1, 702) == "ZZ1"

    def test_triple_letter_columns(self):
        assert position_to_spreadsheet_index(1, 703) == "AAA1"
        assert position_to_spreadsheet_index(1, 16384) == "XFD1"

    def test_large_row(self):
        assert position_to_spreadsheet_index(1048576, 1) == "A1048576"

    def test_invalid_row_zero(self):
        with pytest.raises(ValueError):
            position_to_spreadsheet_index(0, 1)

    def test_invalid_row_negative(self):
        with pytest.raises(ValueError):
            position_to_spreadsheet_index(-1, 1)

    def test_invalid_col_zero(self):
        with pytest.raises(ValueError):
            position_to_spreadsheet_index(1, 0)

    def test_invalid_col_negative(self):
        with pytest.raises(ValueError):
            position_to_spreadsheet_index(1, -5)


# ===========================================================================
# Round-trip
# ===========================================================================

class TestRoundTrip:
    @pytest.mark.parametrize("index", [
        "A1", "Z99", "AA1", "AZ52", "BA100", "ZZ1", "XFD1048576",
    ])
    def test_index_roundtrip(self, index):
        row, col = spreadsheet_index_to_position(index)
        assert position_to_spreadsheet_index(row, col) == index.upper()

    @pytest.mark.parametrize("row,col", [
        (1, 1), (1, 26), (1, 27), (1, 702), (1, 703), (999, 16384),
    ])
    def test_position_roundtrip(self, row, col):
        idx = position_to_spreadsheet_index(row, col)
        assert spreadsheet_index_to_position(idx) == (row, col)


# ===========================================================================
# import_sheets
# ===========================================================================

class TestImportSheets:
    def test_two_sheets(self, sample_xlsx):
        assert import_sheets(sample_xlsx) == ["Data", "Summary"]

    def test_single_sheet(self, empty_xlsx):
        assert import_sheets(empty_xlsx) == ["Empty"]

    def test_five_sheets(self, multi_sheet_xlsx):
        assert import_sheets(multi_sheet_xlsx) == [
            "Alpha", "Beta", "Gamma", "Delta", "Epsilon",
        ]

    def test_accepts_string_path(self, sample_xlsx):
        assert import_sheets(str(sample_xlsx)) == ["Data", "Summary"]


# ===========================================================================
# import_cells
# ===========================================================================

class TestImportCells:
    def test_all_cells_default_sheet(self, sample_xlsx):
        data = import_cells(sample_xlsx)
        assert len(data) == 3
        assert data[0][0] == 10
        assert data[0][1] == 20
        assert data[1][0] == "hello"
        assert data[1][1] == "world"
        assert data[2][0] == 3.14
        assert data[2][1] is True

    def test_formula_cells_return_values(self, sample_xlsx):
        data = import_cells(sample_xlsx)
        # With data_only=True, openpyxl returns cached values (or None if not cached)
        # Formula cells: C1 = =A1+B1, C3 = =SUM(A1:B1)
        # Cached values may be None for freshly created files
        assert data[0][2] is None or data[0][2] == 30

    def test_select_sheet_by_name(self, sample_xlsx):
        data = import_cells(sample_xlsx, sheet="Summary")
        assert data[0][0] == "total"

    def test_select_sheet_by_index(self, sample_xlsx):
        data = import_cells(sample_xlsx, sheet=2)
        assert data[0][0] == "total"

    def test_invalid_sheet_name(self, sample_xlsx):
        with pytest.raises(ValueError, match="not found"):
            import_cells(sample_xlsx, sheet="NoSuchSheet")

    def test_invalid_sheet_index(self, sample_xlsx):
        with pytest.raises(ValueError, match="out of range"):
            import_cells(sample_xlsx, sheet=99)

    def test_row_range(self, sample_xlsx):
        data = import_cells(sample_xlsx, rows=(2, 3))
        assert len(data) == 2
        assert data[0][0] == "hello"
        assert data[1][0] == 3.14

    def test_column_range(self, sample_xlsx):
        data = import_cells(sample_xlsx, columns=(1, 2))
        assert len(data) == 3
        assert all(len(row) == 2 for row in data)
        assert data[0] == [10, 20]

    def test_row_and_column_range(self, sample_xlsx):
        data = import_cells(sample_xlsx, rows=(1, 2), columns=(1, 2))
        assert data == [[10, 20], ["hello", "world"]]

    def test_single_cell_range(self, sample_xlsx):
        data = import_cells(sample_xlsx, rows=(2, 2), columns=(1, 1))
        assert data == [["hello"]]

    def test_empty_sheet(self, empty_xlsx):
        data = import_cells(empty_xlsx)
        assert data == [[None]]

    def test_accepts_string_path(self, sample_xlsx):
        data = import_cells(str(sample_xlsx))
        assert data[0][0] == 10

    def test_accepts_path_object(self, sample_xlsx):
        data = import_cells(Path(sample_xlsx))
        assert data[0][0] == 10


# ===========================================================================
# import_formulas
# ===========================================================================

class TestImportFormulas:
    def test_formula_cells_return_formulas(self, sample_xlsx):
        data = import_formulas(sample_xlsx)
        assert data[0][2] == "=A1+B1"
        assert data[2][2] == "=SUM(A1:B1)"

    def test_non_formula_cells_return_values(self, sample_xlsx):
        data = import_formulas(sample_xlsx)
        assert data[0][0] == 10
        assert data[0][1] == 20
        assert data[1][0] == "hello"

    def test_second_sheet_formula(self, sample_xlsx):
        data = import_formulas(sample_xlsx, sheet="Summary")
        assert data[0][0] == "total"
        assert data[0][1] == "=Data!A1+Data!B1"

    def test_select_sheet_by_index(self, sample_xlsx):
        data = import_formulas(sample_xlsx, sheet=2)
        assert data[0][1] == "=Data!A1+Data!B1"

    def test_row_range(self, sample_xlsx):
        data = import_formulas(sample_xlsx, rows=(1, 1))
        assert len(data) == 1
        assert data[0][2] == "=A1+B1"

    def test_column_range(self, sample_xlsx):
        data = import_formulas(sample_xlsx, columns=(3, 3))
        assert data[0] == ["=A1+B1"]
        assert data[1] == [None]
        assert data[2] == ["=SUM(A1:B1)"]

    def test_empty_sheet(self, empty_xlsx):
        data = import_formulas(empty_xlsx)
        assert data == [[None]]
